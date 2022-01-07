import requests
from urllib import parse
import openpyxl
import time
import common
from histudy import log, request
import json
import os
import specialtest.common as comm


def read_excel_data_toclassin(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    rowend = ws.max_row
    execl_conent_list = ws[2:rowend]
    responsedatalist = []
    for col in execl_conent_list:  # 打印 2-5行中所有单元格中的值
        import requests
        if isinstance(col,tuple):
            course_id = int(col[0].value)
        else:
            course_id = int(col.value)
        # YDD --->Mini班
        print(course_id)
        url = "http://erp.histudy.com/erp/2ta/erp2ta/classin/course/bjk/add"

        payload = f'{{"courseId":{course_id}}}'
        headers = {
            'content-type': "application/json",
            'cookie': common.get_sesion(),
            'cache-control': "no-cache",
            'postman-token': "bb63c29d-abf9-22f1-eb1b-d4003e9ff8b2"
        }

        response = requests.request("POST", url, data=payload, headers=headers)

        print(response.text)
        responsedata = json.loads(response.text)
        responsedatalist.append(responsedata)
        log.info(f"同步mini班课程到classin：" + str(responsedata))
        time.sleep(0.05)  # 间隔多久刷一次数据
    for i in range(2, rowend + 1):  # 2开始
        ws.cell(row=i, column=5).value = str(responsedatalist[i - 2])
    wb.save(filename=filename)
    wb.close()


if __name__ == '__main__':
    paths = os.path.dirname(os.path.abspath(__file__)) + r'\高中课程-待同步classin-福厦泉个性化-0914-毛老师.xlsx'
    print(paths)
    read_excel_data_toclassin(paths)
