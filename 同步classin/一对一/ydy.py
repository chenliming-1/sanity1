#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# !@Author: 练素琼
# !Email: lsqiong_1@histudy.com
# !Date: 2021-09-13 16:41

# 备注：1对1的课程是要获取tab_attend_info的id，为课次id进行同步。

import requests
import openpyxl
import time
import common
from histudy import log
import json
import os

url = "http://erp.histudy.com/erp/2ta/erp2ta/classin/course/ydy/add"


def read_excel_data_toclassin(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    rowend = ws.max_row
    execl_conent_list = ws[2:rowend]
    responsedatalist = []
    for col in execl_conent_list:  # 打印 2-5行中所有单元格中的值
        if isinstance(col,tuple):
            zero_cols_value = int(col[0].value)
        else:
            zero_cols_value = int(col.value)
        # YDD --->Mini班
        print(zero_cols_value)

        payload = f'{{"courseId":{zero_cols_value}}}'
        headers = {
            'content-type': "application/json",
            'cookie': common.get_sesion(),
            'cache-control': "no-cache",
            'postman-token': "bb63c29d-abf9-22f1-eb1b-d4003e9ff8b2"
        }

        response = requests.request("POST", url, data=payload, headers=headers)

        print(response.text)
        responsedata = json.loads(response.text)
        responsedatalist.append(responsedata)
        log.info(f"同步一对一的代码到到classin：" + str(responsedata))
        time.sleep(0.05)  # 间隔多久刷一次数据
    for i in range(2, rowend + 1):  # 2开始
        ws.cell(row=i, column=5).value = str(responsedatalist[i - 2])
    wb.save(filename=filename)
    wb.close()


if __name__ == '__main__':
    paths = os.path.dirname(os.path.abspath(__file__)) + r'\一对一同步classin20210913.xlsx'
    print(paths)
    read_excel_data_toclassin(paths)
