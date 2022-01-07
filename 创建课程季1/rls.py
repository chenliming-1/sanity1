#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# !@Author: 练素琼
# !Email: lsqiong_1@histudy.com
# !Date: 2020-11-03 10:42
import openpyxl
from commidware.recognize import driver
from histudy import request, log, dao
import json
import requests
from pypinyin import pinyin
from datetime import datetime
import sys,os

class CreateSeason():
    def __init__(self):
        self.area_name = None
        self.bu_name = None
        self.branch_name = None
        self.cookie = None
        self.year = None
        self.seanson_name = None
        self.season = None,
        self.season_chinese = None
        self.start_date = None
        self.end_date = None
        self.product_line_name = None
        self.encoding = None
        self.pre_season_name = None,
        self.current_season = None
        self.pre_season = None
        self.product_line = None
        self.season_name_list = ["寒假", "春季", "暑假", "秋季"]
        self.pre_year = None
        self.current_year = None
        # self.env =sys.argv[1]
        # self.url = "http://erp.histudy.com/"
        self.url = "http://erp-aliyun.rls.klxuexi.net/"
        self.driver, self.cookie = driver.get_login_driver("lsqiong_1", "1qazxsw2",
                                                           self.url)

    def readexcel(self, name):
        wb = openpyxl.load_workbook(name)
        # 从表单中获取单元格的内容
        ws = wb.active  # 当前活跃的表单
        rowstart = 2
        rowend = ws.max_row
        execl_conent_list = ws[rowstart:rowend]
        res_json = {}
        res_list = []
        for item in execl_conent_list:
            self.area_name = item[0].value
            self.bu_name = item[1].value
            self.branch_name = item[2].value
            res_list.append([self.area_name, self.bu_name, self.branch_name])
        return res_list

    def change_org(self, branch_name):
        try:
            branch_id = \
                dao(f'erp_rls', f"select id from tab_organization_info where org_name='{branch_name}'")[0]["ID"]
            log.info("运行切换校区接口，切换至校区为：{0}的校区".format(branch_name))
            self.changeorg_rsp = request.run_main(
                f"{self.url}/common/orgservice?id=%d" % int(branch_id), 'PUT',
                data={},
                headers={
                    "Cookie": self.cookie,
                    "Accept": "application/json, text/plain, */*",
                    "Content-Type": "application/json;charset=UTF-8"})
            self.rsp_error = self.changeorg_rsp.json()['error']
            self.rsp_changeTeam = self.changeorg_rsp.json()['changeTeam']
        except Exception as ex:
            log.error("切换校区异常，失败原因为：{0}".format(ex))
        finally:
            return self.rsp_changeTeam

    def gen_season_name(self, area_name, product_name, season_name):
        now_date = datetime.now().strftime('%Y')
        now_date =int(now_date) + 1
        self.seanson_name = str(now_date) + "年" + area_name + product_name + season_name + "班"
        print(self.seanson_name)
        return self.seanson_name

    def init_data(self, season_name, encoding, year, seanson, pre_season_name, pre_season_id, start_date, end_date):
        init_data = {
            "business_type": 1,  # 新增的时候是不是都是培英班类型的课程季
            "business_type_name": "",
            "city_id": 0,
            "city_name": "",
            "course_season_name": season_name,
            "create_time": None,
            "create_user": 0,
            "create_user_name": "",
            "description": "",
            "end_date": end_date,
            "id": 0,
            "last_course_season_name": pre_season_name,
            "last_season_id": pre_season_id,  # 上一课程季
            "product_line": None,
            "product_line_name": "",
            "season": seanson,
            "season_name": "",
            "start_date": start_date,
            "status": 0,
            "update_time": None,
            "update_user": 0,
            "update_user_name": "",
            "years": year,
            "encoding": encoding
        }
        return init_data

    def add_season(self, filename):
        headers = {
            "Cookie": self.cookie,
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8"
        }
        result = self.readexcel(filename)
        item = [item for item in result]
        for i in item:
            self.area_name = i[0]
            self.bu_name = i[1]
            self.branch_name = i[2]
            self.change_org(self.branch_name)  # 切换到校区
            self.product_line = \
                dao("erp_rls", f"select product_line from tab_organization_info where org_name ='{self.branch_name}'")[
                    0][
                    "PRODUCT_LINE"]
            if  self.product_line == 11:
                self.product_line_name = "佳音"
            elif self.product_line == 1:
                self.product_line_name = "培英班"
            elif self.product_line == 2:
                self.product_line_name = "个性化"
            elif self.product_line == 12:
                self.product_line_name = "双师"
            for season_name in self.season_name_list:
                self.seanson_name = self.gen_season_name(self.area_name, self.product_line_name, season_name)
                self.encoding = self.getFirstPinYin(self.seanson_name)
                pre_season = self.count_pre_season(season_name)
                print(pre_season[2])  # 是否要通过,分割生成一个数组
                self.pre_season_name = str(pre_season[6]) + "年" + self.area_name + self.product_line_name + pre_season[
                    3] + "班"
                print(self.pre_season_name)
                print(f"SELECT tts.id FROM tab_time_season tts left join bu_dict_rel bdr on bdr.dict_id = tts.id left join tab_organization_info toi on toi.id = bdr.bu_id where org_name='{self.bu_name}' and tts.years='{pre_season[6]}' and tts.season ='{pre_season[2]}' and tts.status ='1'")
                # 判断是否已经存在了
                pre_season_id = dao("erp_rls",
                                    f"SELECT tts.id FROM tab_time_season tts left join bu_dict_rel bdr on bdr.dict_id = tts.id left join tab_organization_info toi on toi.id = bdr.bu_id where org_name='{self.bu_name}' and tts.years='{pre_season[6]}' and tts.season ='{pre_season[2]}' and tts.status ='1' ")
                print(pre_season_id)
                if pre_season_id:
                    pre_season_id = dao("erp_rls",
                                        f"SELECT tts.id FROM tab_time_season tts left join bu_dict_rel bdr on bdr.dict_id = tts.id left join tab_organization_info toi on toi.id = bdr.bu_id where toi.org_name='{self.bu_name}' and tts.years='{pre_season[6]}' and tts.season ='{pre_season[2]}' and tts.status ='1' ")[0]["ID"]
                else:
                    pre_season_id = -1
                now_date = datetime.now().strftime('%Y')
                current_year=str(int(now_date)+1)
                # 判断years和seasom是否已经存在，是判断当下的课程季是否已经有了，不能取到上一个课程季的值进行判断
                years = \
                    dao("erp_rls",
                        f"SELECT tts.years FROM tab_time_season tts left join bu_dict_rel bdr on bdr.dict_id = tts.id left join tab_organization_info toi on toi.id = bdr.bu_id where  toi.org_name='{self.bu_name}' and tts.years='{current_year}' and tts.season ='{pre_season[1]}' and tts.status ='1' ")

                seasom = \
                    dao("erp_rls",
                        f"SELECT tts.season FROM tab_time_season tts left join bu_dict_rel bdr on bdr.dict_id = tts.id left join tab_organization_info toi on toi.id = bdr.bu_id where toi.org_name='{self.bu_name}'  and tts.years='{current_year}' and tts.season ='{pre_season[1]}' and tts.status ='1' ")
                if years and seasom: # 查出来已经有值，不为空，true,需要pass,只有课程季不存在的时候才要写 （这个查询出来是上一个课程季是否有，不是当前课程季是否有错了）
                    pass
                else:
                    request_json = self.init_data(self.seanson_name, self.encoding,  int(now_date)+1, pre_season[1],
                                                  self.pre_season_name, pre_season_id, pre_season[4], pre_season[5])

                    self.response=request.run_main(self.url+"erp/dictionary/timeSeason/service",
                                     "POST",data=request_json,headers=headers)
                    print(self.response.text)

    def getFirstPinYin(self, season_name):
        season_name = pinyin(season_name)
        listaa = []
        for item in season_name:
            for i in item:
                if i.isdecimal():
                    listaa.append(i)
                else:
                    first = i[0:1]
                    listaa.append(first)
        str1 = ''.join(listaa)
        return str1

    def count_pre_season(self, season):
        now_date = datetime.now().strftime('%Y')
        self.current_year = int(now_date)+1
        if season in "寒假":
            self.pre_year = int(now_date)  # 计算出冬季的年需要减一 上一年
            self.current_season = "4"
            self.pre_season = "3"
            # self.season = int(self.season) - 1  # 课程季要减一
            self.season_chinese = "秋季"
            self.start_date = str(self.current_year) + "-01-01"
            self.end_date = str(self.current_year) + "-02-28"
        elif season in "秋季":
            self.pre_year = self.current_year
            self.current_season = "3"
            self.pre_season = "2"
            self.season_chinese = "暑假"
            self.start_date = str(self.current_year) + "-09-01"
            self.end_date = str(self.current_year) + "-12-31"

        elif season in "暑假":
            self.pre_year = self.current_year
            self.current_season = "2"
            self.pre_season = "1"
            self.season_chinese = "春季"
            self.start_date = str(self.current_year) + "-07-01"
            self.end_date = str(self.current_year) + "-08-31"

        elif season in "春季":
            self.pre_year = self.current_year
            self.current_season = "1"
            self.pre_season = "4"
            self.season_chinese = "寒假"
            self.start_date = str(self.current_year) + "-03-01"
            self.end_date = str(self.current_year) + "-06-30"
        return self.current_year, self.current_season,self.pre_season, self.season_chinese, self.start_date, self.end_date, self.pre_year

        # return f"{self.year}, {self.season}, 上一个课程季：{self.season_chinese},当前课程季{season}的开始和结束时间： {self.start_date}, {self.end_date}"  # 用于拼接上一个课程季的数据




if __name__ == '__main__':
    season = CreateSeason()
    import os

    # # 切换校区
    # # season.change_org("北京鸟巢双师校区（主场）")
    #
    # season.add_season()
    # current_path = os.path.abspath(os.path.dirname(__file__))
    # aa = season.readexcel(current_path + "/rls课程季.xlsx")
    # for item in aa:
    #     for i in item:
    #         print(i)
    # dao("erp_rls","select * from ")

    # pinyin=pinyin("厦门测试")
    # print(season.getFirstPinYin(pinyin))
    # list = season.gen_season_name("rls课程季.xlsx", "培英", "冬季")
    # for i in list:
    #     print(i)
    #     encoding = season.getFirstPinYin(i)
    #
    # pre_season = season.count_pre_season("冬季")
    # print(pre_season)
    current_path = os.path.abspath(os.path.dirname(__file__))
    season.add_season(current_path + "/rls课程季.xlsx")
