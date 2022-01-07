import requests
from urllib import parse
import openpyxl
import time
import common
from histudy import log, request
import json


def read_excel_data_toclassin(filename, rowstart, rowend):
    wb = openpyxl.load_workbook(filename)
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    col_range = ws[rowstart:rowend]
    responselist = []
    responsedatalist = []
    for col in col_range:  # 打印 2-5行中所有单元格中的值
        attend_id = col[0].value
        course_id = int(attend_id)
        # course_id = int(100233695)
        print(course_id)
        import requests
        import requests
        # YDD --->Mini班
        url = f"https://erp.histudy.com//erp/syncMtErpData2MtSystem/syncCourseInfo?isMtCourse=Y&"    ###courseId={course_id}这个要获取到主场的课程ID？

        # payload = f'{{"courseId":{course_id}}}'
        headers = {
            'content-type': "application/json",
            'cookie': common.get_sesion(),
            'cache-control': "no-cache",
            'postman-token': "bb63c29d-abf9-22f1-eb1b-d4003e9ff8b2"
        }
        response = requests.request("GET", url, data={}, headers=headers)

        print(response.text)
        responsedata = json.loads(response.text)
        responsedatalist.append(responsedata)
        log.info(f"同步课程成功" + str(responsedata))
        time.sleep(1)  # 间隔多久刷一次数据
    for i in range(rowstart, rowend + 1):  # 2开始
        ws.cell(row=i, column=10).value = str(responsedatalist[i - 2])
    wb.save(filename=filename)


if __name__ == '__main__':
    read_excel_data_toclassin(r"课程教材处理数据.xlsx", 237, 288)
