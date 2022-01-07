#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# !@Author: 练素琼
# !Email: lsqiong_1@histudy.com
# !Date: 2021-06-25 9:49
import requests
import openpyxl
from specialtest.common import get_sesion
from histudy import request
import json


class CancleScheduling():
    def __init__(self):
        self.hostname = "erp-manager.histudy.com"
        self.url = "http://{}/api/attendance/ydyBatchAttendAndCheck".format(self.hostname)
        self.header = {
            "Cookie": str(get_sesion()),
            "Content-Type": "application/json"
        }

        self.data = [{
            "attendId": None,
            "attendType": "YDY_QX",
            "subAttendType": "YDY_QX",
            "remark": "系统处理排课取消-支持半小时耗课",
            "source": 5
        }]

    def cancle(self, filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        rowend = ws.max_row
        execl_conent_list = ws[2:rowend]
        resultlist = []
        for item in execl_conent_list:
            self.attend_id = item[0].value
            self.data[0]["attendId"] = self.attend_id
            self.result = request.run_main(self.url, method="POST", headers=self.header, data=self.data)
            print(self.result.text)
            resultlist.append(self.result.text)
        for i in range(2, rowend + 1):  # 2开始
            ws.cell(row=i, column=2).value = str(resultlist[i - 2])
        wb.save(filename=filename)
        wb.close()


if __name__ == '__main__':
    cancel = CancleScheduling()
    import os
    paths = os.path.dirname(os.path.abspath(__file__)) + r'\剩余存在取消排课记录.xlsx'
    cancel.cancle(paths)
