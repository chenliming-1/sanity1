import openpyxl
from histudy import request, log
from specialtest.common import get_sesion
import requests
import  time
import os
import json
#创建批量考勤数据类
class attendData():
    def __init__(self):
        self.hostname = "erp.test.klxuexi.net"
        self.url ="http://{}/erp/attendance/students".format(self.hostname)
        self.header = {
            "Cookie":str(get_sesion()),
            "Accept": "application/json, text/plain, */*",
            "Conten-Type":"application/json;charset=UTF-8",
            'cache-control': "no-cache"
        }
        print("打印请求头：",self.header)
    #读取表格数据
    def read_excel_attend_data(self,filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        rowend = ws.max_row
        colmax = ws.max_column
        # execl_conent_list = ws[2:rowend]
        # print("excellist",execl_conent_list)
        # results = []
        #获取表格的数据为列表[(),()]
        all_data = list(ws.values)
        print("所有的数据：",all_data)
        #获取从表格第二行开始的每一行数据，并且用参数表示
        for row in range(1,rowend):
            print(row)
            data = all_data[row]
            # print(data)
            data_schedulingId = data[0]
            data_studentId = data[1]
            data_studentName = data[2]
            data_lock_status = data[3]
            data_teacherId=data[4]
            data_remark = data[5]
            data_order_encoding = data[6]
            data_courseDate = data[7]
            data_attendType = data[8]
            data_respon = data[9]
            #请求体
            payload = {"submitAttendanceList":[{
                "schedulingId":data_schedulingId,
                "studentId":data_studentId,
                "studentName":data_studentName,
                "lock_status":data_lock_status,
                "teacherId":data_teacherId,
                "remark":data_remark,
                "order_encoding":data_order_encoding,
                "courseDate":data_courseDate,
                "attendType":data_attendType,
            }]}

            # 打印请求的数据
            print("打印请求的数据:",payload)
            #发起一个请求
            result=requests.post(url=self.url,headers=self.header,json=payload)
            # result = request.run_main(url=self.url,method="POST",headers=self.header,data=json)
            print("打印响应状态码：",result.status_code)
            print("打印响应数据：",result.text)
            print("等待数据执行...")
            #当前时间
            # t = time.asctime()
            # print("接口执行的当前时间：",t)
            # 获取接口响应时间
            #计算的是从发送请求到服务端响应回来这段时间(也就是时间差)，发送第一个数据到收到最后一个数据之间，这个时长不受响应的内容影响
            print("接口响应时间：",result.elapsed)
            print("接口响应时间：", result.elapsed.total_seconds())

        # results.append(all_data())



if __name__ == '__main__':
    paths = os.path.dirname(os.path.abspath(__file__)) + r'\批量置空考勤.xlsx'
    print(paths)
    atd = attendData()
    atd.read_excel_attend_data(paths)








#打开工作簿--找到页面--按行读取每一行信息--存到列表当中